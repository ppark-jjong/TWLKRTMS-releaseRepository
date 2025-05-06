"""
엑셀 내보내기 기능 관련 라우터 - 관리자 전용
"""

from typing import Dict, Any, List, Optional
from datetime import datetime, date
import io
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
import logging

from main.utils.database import get_db
from main.utils.security import get_admin_user
from main.models.dashboard_model import Dashboard
from main.service.dashboard_service import get_dashboard_list

# 로깅 설정
logger = logging.getLogger(__name__)

# 라우터 생성 (관리자 전용)
api_router = APIRouter(prefix="/api", dependencies=[Depends(get_admin_user)])

@api_router.get("/dashboard/export-excel")
async def export_orders_to_excel(
    db: Session = Depends(get_db),
    start_date: Optional[date] = Query(None, description="조회 시작일"),
    end_date: Optional[date] = Query(None, description="조회 종료일"),
    current_user: Dict[str, Any] = Depends(get_admin_user),  # 관리자 권한 필요
):
    """주문 목록을 엑셀 파일로 내보내는 API (관리자 전용)"""
    logger.info(f"주문 엑셀 다운로드 요청: user={current_user.get('user_id')}")
    try:
        # 날짜 범위 처리
        today = datetime.now().date()
        final_start_date = start_date or today
        final_end_date = end_date or today

        if final_start_date > final_end_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="시작일이 종료일보다 클 수 없습니다."
            )

        # 서비스 함수로 데이터 조회
        orders = get_dashboard_list(
            db=db, start_date=final_start_date, end_date=final_end_date
        )
        
        if not orders:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="해당 기간에 주문 데이터가 없습니다."
            )

        # 데이터를 DataFrame으로 변환
        orders_data = []
        for order in orders:
            # 주문 상태 한글 라벨 매핑
            status_label = {
                "WAITING": "대기",
                "IN_PROGRESS": "진행",
                "COMPLETE": "완료",
                "ISSUE": "이슈",
                "CANCEL": "취소",
            }.get(order.status, order.status)
            
            # 타입 한글 라벨 매핑
            type_label = {
                "DELIVERY": "배송",
                "RETURN": "회수"
            }.get(order.type, order.type)
            
            # 각 주문 데이터를 딕셔너리로 변환
            row = {
                "주문번호": order.order_no,
                "유형": type_label,
                "상태": status_label,
                "부서": order.department,
                "창고": order.warehouse,
                "SLA": order.sla,
                "ETA": order.eta.strftime("%Y-%m-%d %H:%M") if order.eta else "",
                "생성시간": order.create_time.strftime("%Y-%m-%d %H:%M") if order.create_time else "",
                "출발시간": order.depart_time.strftime("%Y-%m-%d %H:%M") if order.depart_time else "",
                "완료시간": order.complete_time.strftime("%Y-%m-%d %H:%M") if order.complete_time else "",
                "우편번호": order.postal_code,
                "지역": order.region or "",
                "주소": order.address,
                "고객명": order.customer,
                "연락처": order.contact or "",
                "기사명": order.driver_name or "",
                "기사연락처": order.driver_contact or "",
                "거리(km)": order.distance or "",
                "소요시간(분)": order.duration_time or "",
                "메모": order.remark or "",
                "최종수정자": order.update_by or "",
                "최종수정시간": order.update_at.strftime("%Y-%m-%d %H:%M") if order.update_at else "",
            }
            orders_data.append(row)

        # DataFrame 생성
        df = pd.DataFrame(orders_data)

        # 엑셀 파일 생성
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='주문 목록', index=False)
            
            # 워크시트 가져오기
            workbook = writer.book
            worksheet = writer.sheets['주문 목록']
            
            # 열 너비 조정
            for i, column in enumerate(df.columns):
                column_width = max(df[column].astype(str).map(len).max(), len(column) + 2)
                # 최소 너비 8, 최대 너비 50으로 제한
                column_width = min(max(column_width, 8), 50)
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = column_width
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        # 파일 포인터를 처음으로 되돌림
        output.seek(0)
        
        # 파일명 생성 (현재 날짜 기준)
        today_str = datetime.now().strftime("%Y%m%d")
        filename = f"주문목록_{today_str}.xlsx"
        
        # 파일 응답 생성
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException as http_exc:
        # 이미 HTTPException이면 그대로 발생
        logger.warning(f"엑셀 다운로드 HTTP 오류: {http_exc.detail}")
        raise http_exc
    except Exception as e:
        logger.error(f"엑셀 파일 생성 중 오류: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="엑셀 파일 생성 중 오류가 발생했습니다."
        )
