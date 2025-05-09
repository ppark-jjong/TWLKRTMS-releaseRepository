"""
대시보드(주문) 관련 라우터 - 리팩토링 버전
"""

from typing import Dict, Any, List, Optional, Union, Tuple
from datetime import datetime, date
from decimal import Decimal
import json
import logging
import sys
import traceback
from urllib.parse import quote
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Request,
    Query,
    Path,
    status,
    Form,
    Body,
)
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError

from main.core.templating import templates
from main.utils.database import get_db, db_transaction
from main.utils.security import get_current_user, get_admin_user
from main.models.dashboard_model import Dashboard
from main.schema.dashboard_schema import (
    DashboardCreate,
    DashboardUpdate,
    DashboardResponse,
    DashboardListResponse,
    DashboardDeleteRequest,
)
from main.service.dashboard_service import (
    get_dashboard_by_id,
    get_dashboard_list,
    search_dashboard_by_order_no,
    create_dashboard,
    update_dashboard,
    change_status,
    assign_driver,
    delete_dashboard,
    get_dashboard_list_paginated,
    get_dashboard_response_data,
    get_dashboard_list_item_data,
    get_dashboard_by_order_no,
)
from main.utils.json_util import CustomJSONEncoder
from pydantic import BaseModel  # Pydantic 모델 사용 위해 추가

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        # 파일 핸들러는 필요시 활성화
        # logging.FileHandler("dashboard_route.log"),
    ],
)
logger = logging.getLogger(__name__)

# 라우터 생성
api_router = APIRouter(prefix="/api", dependencies=[Depends(get_current_user)])
page_router = APIRouter(dependencies=[Depends(get_current_user)])

# --- 상태 전이 규칙 (JS와 동일하게 유지 - 재정의된 규칙) ---
status_transitions = {  # 일반 사용자
    "WAITING": ["IN_PROGRESS"],
    "IN_PROGRESS": ["COMPLETE", "ISSUE", "CANCEL"],
    "COMPLETE": ["ISSUE", "CANCEL"],
    "ISSUE": ["COMPLETE", "CANCEL"],
    "CANCEL": ["COMPLETE", "ISSUE"],
}
admin_status_transitions = {  # 관리자
    "WAITING": ["IN_PROGRESS"],
    "IN_PROGRESS": ["WAITING", "COMPLETE", "ISSUE", "CANCEL"],
    "COMPLETE": ["IN_PROGRESS", "ISSUE", "CANCEL"],
    "ISSUE": ["IN_PROGRESS", "COMPLETE", "CANCEL"],
    "CANCEL": ["IN_PROGRESS", "COMPLETE", "ISSUE"],
}

# --- 상태 라벨 (JS와 공유 또는 백엔드에서만 사용) ---
status_labels = {
    "WAITING": "대기",
    "IN_PROGRESS": "진행",
    "COMPLETE": "완료",
    "ISSUE": "이슈",
    "CANCEL": "취소",
}


def get_next_possible_statuses(current_status: str, is_admin: bool) -> List[str]:
    """현재 상태와 사용자 역할에 따라 다음 가능한 상태 목록 반환 (재정의된 규칙 사용)"""
    if is_admin:
        return admin_status_transitions.get(current_status, [])
    else:
        return status_transitions.get(current_status, [])


# === 유틸리티 함수 ===
def handle_redirect_error(
    error: Union[Exception, str],
    redirect_url: str,
    status_code: int = status.HTTP_303_SEE_OTHER,
) -> RedirectResponse:
    """오류를 처리하고 리다이렉션 응답을 생성하는 헬퍼 함수"""
    if isinstance(error, Exception):
        error_message = getattr(error, "detail", str(error))
    else:
        error_message = error

    # 메시지 인코딩 및 리다이렉션
    encoded_message = quote(error_message)
    return RedirectResponse(
        f"{redirect_url}?error={encoded_message}", status_code=status_code
    )


def handle_redirect_success(
    message: str, redirect_url: str, status_code: int = status.HTTP_303_SEE_OTHER
) -> RedirectResponse:
    """성공 메시지를 처리하고 리다이렉션 응답을 생성하는 헬퍼 함수"""
    encoded_message = quote(message)
    return RedirectResponse(
        f"{redirect_url}?success={encoded_message}", status_code=status_code
    )


# === 페이지 렌더링 라우트 ===
@page_router.get("/dashboard", include_in_schema=False, name="dashboard_page")
async def get_dashboard_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    logger.info(f"대시보드 페이지 요청: user={current_user.get('user_id')}")
    try:
        error_message = request.query_params.get("error")
        success_message = request.query_params.get("success")

        user_data = {
            "user_id": current_user.get("user_id"),
            "user_role": current_user.get("user_role"),
            "department": current_user.get("department"),
        }

        context = {
            "request": request,
            "current_user": user_data,
            "error_message": error_message,
            "success_message": success_message,
        }

        return templates.TemplateResponse("dashboard.html", context)

    except HTTPException as http_exc:
        logger.error(
            f"대시보드 페이지 로드 중 HTTP 오류: {http_exc.detail}", exc_info=True
        )
        context = {
            "request": request,
            "error_message": http_exc.detail,
            "current_user": current_user,
        }
        return templates.TemplateResponse(
            "error.html", context, status_code=http_exc.status_code
        )
    except Exception as e:
        logger.error(f"대시보드 페이지 렌더링 중 예외 발생: {e}", exc_info=True)
        context = {
            "request": request,
            "error_message": "페이지 로드 중 오류 발생",
            "current_user": current_user,
        }
        return templates.TemplateResponse("error.html", context, status_code=500)


# === API 엔드포인트 라우트 ===
@api_router.post("/orders/batch-update")
@db_transaction
async def batch_update_orders(
    request: Request,
    update_data: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    주문 일괄 변경 API
    여러 주문의 상태나 기사 정보를 한 번에 변경합니다.
    """
    user_id = current_user.get("user_id")
    user_role = current_user.get("user_role")
    logger.info(f"주문 일괄 변경 API 요청: user={user_id}")

    try:
        # 요청 데이터 검증
        if not update_data or "ids" not in update_data or not update_data.get("update"):
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "message": "잘못된 요청 형식입니다."},
            )

        order_ids = update_data.get("ids", [])
        update_fields = update_data.get("update", {})

        if not order_ids:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "message": "변경할 주문을 선택해주세요."},
            )

        # 상태 변경 또는 기사 배정 처리
        results = []
        succeeded = 0

        if "status" in update_fields:
            # 상태 일괄 변경
            new_status = update_fields["status"]
            status_results = change_status(
                db=db,
                dashboard_ids=order_ids,
                new_status=new_status,
                user_id=user_id,
                user_role=user_role,
            )

            # 결과 처리
            for result in status_results:
                if result.get("success"):
                    succeeded += 1
                results.append(result)

        elif "driver_name" in update_fields:
            # 기사 일괄 배정
            driver_name = update_fields.get("driver_name")
            driver_contact = update_fields.get("driver_contact")

            assign_results = assign_driver(
                db=db,
                dashboard_ids=order_ids,
                driver_name=driver_name,
                driver_contact=driver_contact,
                user_id=user_id,
            )

            # 결과 처리
            for result in assign_results:
                if result.get("success"):
                    succeeded += 1
                results.append(result)
        else:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "message": "지원하지 않는 변경 유형입니다."},
            )

        # 성공 응답
        return JSONResponse(
            content={
                "success": True,
                "message": f"{succeeded}개 주문이 성공적으로 변경되었습니다.",
                "succeeded": succeeded,
                "failed": len(order_ids) - succeeded,
                "results": results,
            }
        )

    except Exception as e:
        logger.error(f"주문 일괄 변경 처리 중 오류: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={
                "success": False,
                "message": f"일괄 변경 중 오류가 발생했습니다: {str(e)}",
            },
        )


@api_router.get("/dashboard/list", response_model=DashboardListResponse)
async def get_dashboard_list_api(
    db: Session = Depends(get_db),
    start_date: Optional[date] = Query(None, description="조회 시작일"),
    end_date: Optional[date] = Query(None, description="조회 종료일"),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    logger.info(f"대시보드 목록 API 호출: user={current_user.get('user_id')}")
    try:
        today = datetime.now().date()
        final_start_date = start_date or today
        final_end_date = end_date or today

        if final_start_date > final_end_date:
            return DashboardListResponse(
                success=False, message="잘못된 날짜 범위입니다.", data=[]
            )

        orders_raw = get_dashboard_list(
            db=db, start_date=final_start_date, end_date=final_end_date
        )
        orders_data = [get_dashboard_list_item_data(order) for order in orders_raw]

        return DashboardListResponse(
            success=True, message="주문 목록 조회 성공", data=orders_data
        )

    except HTTPException as http_exc:
        # 서비스 레벨에서 발생한 HTTPException 처리
        logger.warning(f"목록 조회 API 오류 (HTTPException): {http_exc.detail}")
        raise http_exc
    except Exception as e:
        logger.error(f"목록 조회 API 오류: {e}", exc_info=True)
        # Pydantic 검증 실패 포함 가능
        error_detail = str(e)
        if "ValidationError" in error_detail:
            error_detail = "데이터 형식 검증에 실패했습니다."  # 사용자 친화적 메시지

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"목록 조회 중 오류 발생: {error_detail}",
        )


@api_router.get("/dashboard/export-excel")
async def export_dashboard_to_excel(
    start_date: Optional[date] = Query(None, description="조회 시작일"),
    end_date: Optional[date] = Query(None, description="조회 종료일"),
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_admin_user),  # 관리자 전용 API
):
    """
    관리자 전용 엑셀 다운로드 API
    지정한 날짜 범위의 주문 데이터를 엑셀 파일로 내보냅니다.
    """
    logger.info(
        f"엑셀 다운로드 요청: user={current_user.get('user_id')}, 시작일={start_date}, 종료일={end_date}"
    )

    try:
        # 날짜 기본값 설정
        today = datetime.now().date()
        final_start_date = start_date or today
        final_end_date = end_date or today

        # 날짜 유효성 검사
        if final_start_date > final_end_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="종료일은 시작일보다 같거나 늦어야 합니다.",
            )

        # 주문 데이터 가져오기
        orders = get_dashboard_list(
            db=db, start_date=final_start_date, end_date=final_end_date
        )

        if not orders:
            df = pd.DataFrame()
        else:
            # 데이터를 Pandas DataFrame으로 변환 (영문 컬럼명 사용)
            order_dicts = []
            for order in orders:
                order_dict = {
                    "order_no": order.order_no,
                    "type": order.type,
                    "department": order.department,
                    "warehouse": order.warehouse,
                    "sla": order.sla,
                    "status": order.status,
                    "eta": order.eta.strftime("%Y-%m-%d %H:%M") if order.eta else None,
                    "create_time": (
                        order.create_time.strftime("%Y-%m-%d %H:%M")
                        if order.create_time
                        else None
                    ),
                    "depart_time": (
                        order.depart_time.strftime("%Y-%m-%d %H:%M")
                        if order.depart_time
                        else None
                    ),
                    "complete_time": (
                        order.complete_time.strftime("%Y-%m-%d %H:%M")
                        if order.complete_time
                        else None
                    ),
                    "postal_code": order.postal_code,
                    "region": order.region or None,
                    "address": order.address,
                    "customer": order.customer,
                    "contact": order.contact or None,
                    "driver_name": order.driver_name or None,
                    "driver_contact": order.driver_contact or None,
                    "remark": order.remark or None,
                    "update_by": order.update_by or None,
                    "update_at": (
                        order.update_at.strftime("%Y-%m-%d %H:%M")
                        if order.update_at
                        else None
                    ),
                }
                order_dicts.append(order_dict)

            df = pd.DataFrame(order_dicts)

        # 엑셀 파일 생성 (openpyxl 직접 사용)
        output = io.BytesIO()
        # Pandas ExcelWriter 대신 openpyxl 직접 사용
        # with pd.ExcelWriter(output, engine="openpyxl") as writer:
        #     df.to_excel(writer, sheet_name="주문목록", index=False)

        # openpyxl을 사용하여 DataFrame을 쓰고 스타일 적용
        df.to_excel(output, index=False, sheet_name="DashboardData")  # 시트 이름 변경
        output.seek(0)  # 중요: to_excel 후 포인터 되돌리기

        # 워크북 로드 및 스타일 적용
        from openpyxl import load_workbook

        wb = load_workbook(output)
        ws = wb.active

        # 헤더 스타일 정의 (볼드 제거, 회색 배경)
        header_font = Font(bold=False)
        header_fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )

        # 첫 번째 행(헤더)에 스타일 적용
        for cell in ws[1]:  # 첫 번째 행의 모든 셀
            cell.font = header_font
            cell.fill = header_fill

        # 수정된 워크북을 다시 BytesIO에 저장
        output.seek(0)
        output.truncate()  # 기존 내용 지우기
        wb.save(output)
        output.seek(0)

        # 파일명 생성 (YYMMDD 형식)
        start_str = final_start_date.strftime("%y%m%d")
        end_str = final_end_date.strftime("%y%m%d")
        filename_base = f"TWLKR-dashboard_{start_str}_{end_str}.xlsx"

        # 파일명 URL 인코딩 (UTF-8)
        filename_encoded = quote(filename_base)

        # 스트리밍 응답 반환 (헤더 수정)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                # RFC 6266 अनुसार Content-Disposition हेडर सेट करें
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"
            },
        )

    except HTTPException as http_exc:
        # HTTP 오류 그대로 전달
        logger.warning(
            f"엑셀 다운로드 HTTP 오류: {http_exc.status_code}, {http_exc.detail}"
        )
        raise http_exc
    except Exception as e:
        # 기타 오류 처리
        logger.error(f"엑셀 다운로드 중 오류 발생: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"엑셀 파일 생성 중 오류가 발생했습니다: {str(e)}",
        )


@api_router.get("/dashboard/search")
async def search_order_api(
    order_no: str = Query(..., description="검색할 주문번호"),
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    logger.info(
        f"주문번호 검색 API 호출: order_no='{order_no}', user={current_user.get('user_id')}"
    )
    order_no_trimmed = order_no.strip()
    if not order_no_trimmed:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "검색할 주문번호를 입력해주세요."},
        )

    try:
        order = search_dashboard_by_order_no(db=db, order_no=order_no_trimmed)
        order_data = get_dashboard_response_data(order) if order else None
        message = (
            f"'{order_no_trimmed}' 검색 결과"
            if order
            else f"'{order_no_trimmed}'에 해당하는 주문 없음"
        )
        return {"success": True, "message": message, "data": {"order": order_data}}

    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        logger.error(f"주문번호 검색 중 오류: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="주문 검색 중 오류가 발생했습니다.",
        )


# === 추가: 일괄 처리 API ===


class BatchStatusUpdateRequest(BaseModel):
    dashboard_ids: List[int]
    new_status: str


@api_router.post(
    "/orders/batch-status-update",
    summary="선택된 주문들의 상태 일괄 변경",
    response_model=List[Dict[str, Any]],
)
@db_transaction
async def batch_update_order_status(
    request: Request,
    update_request: BatchStatusUpdateRequest,
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """선택된 주문 ID 목록을 받아 상태를 일괄 변경합니다."""
    user_id = current_user.get("user_id")
    user_role = current_user.get("user_role")
    logger.info(
        f"상태 일괄 변경 요청: IDs {update_request.dashboard_ids}, NewStatus {update_request.new_status}, User {user_id}"
    )

    try:
        results = change_status(
            db=db,
            dashboard_ids=update_request.dashboard_ids,
            new_status=update_request.new_status,
            user_id=user_id,
            user_role=user_role,
        )
        # 성공/실패 여부에 따라 적절한 상태 코드 반환 고려 (예: 일부만 성공 시 207 Multi-Status)
        # 여기서는 일단 모든 결과를 200 OK로 반환
        return JSONResponse(content=results)
    except HTTPException as http_exc:
        # change_status 내에서 발생하는 권한 오류 등
        logger.warning(f"상태 일괄 변경 중 HTTP 오류: {http_exc.detail}")
        raise http_exc
    except Exception as e:
        logger.error(f"상태 일괄 변경 중 서버 오류: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="상태 일괄 변경 중 서버 오류 발생",
        )


class BatchDriverAssignRequest(BaseModel):
    dashboard_ids: List[int]
    driver_name: str
    driver_contact: Optional[str] = None
    delivery_company: Optional[str] = None


@api_router.post(
    "/orders/batch-driver-assign",
    summary="선택된 주문들에 기사/배송사 일괄 할당",
    response_model=List[Dict[str, Any]],
)
@db_transaction
async def batch_assign_driver_company(
    request: Request,
    assign_request: BatchDriverAssignRequest,
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """선택된 주문 ID 목록을 받아 기사 정보 및 배송사를 일괄 할당합니다."""
    user_id = current_user.get("user_id")
    logger.info(
        f"기사/배송사 일괄 할당 요청: IDs {assign_request.dashboard_ids}, Driver {assign_request.driver_name}, Company {assign_request.delivery_company}, User {user_id}"
    )

    try:
        results = assign_driver(
            db=db,
            dashboard_ids=assign_request.dashboard_ids,
            driver_name=assign_request.driver_name,
            driver_contact=assign_request.driver_contact,
            delivery_company=assign_request.delivery_company,
            user_id=user_id,
        )
        return JSONResponse(content=results)
    except Exception as e:
        logger.error(f"기사/배송사 일괄 할당 중 서버 오류: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="기사/배송사 일괄 할당 중 서버 오류 발생",
        )


# --- 주문 생성 페이지 --- (라우터는 페이지 렌더링에 집중)
@page_router.get("/orders/new", name="order_create_page")
async def order_create_page(
    request: Request,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    logger.info(f"주문 생성 페이지 로드 요청: user={current_user.get('user_id')}")
    try:
        # 설명서에 따라 Form(PRG) 방식 적용 - 템플릿에 직접 객체 전달
        context = {
            "request": request,
            "current_user": current_user,
            "is_edit": False,
            "order": None,
            "error_message": request.query_params.get("error"),
            "success_message": request.query_params.get("success"),
        }
        return templates.TemplateResponse("order_form.html", context)
    except Exception as e:
        logger.error(f"주문 생성 페이지 로드 오류: {e}", exc_info=True)
        # 오류 발생 시 대시보드로 리다이렉트
        error_message = quote("페이지 로드 중 오류 발생")
        return RedirectResponse(
            url=f"/dashboard?error={error_message}",
            status_code=status.HTTP_303_SEE_OTHER,
        )


# --- 주문 상세 페이지 --- (라우터는 페이지 렌더링에 집중)
@page_router.get("/orders/{dashboard_id}", name="order_detail_page")
async def order_detail_page(
    request: Request,
    dashboard_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    logger.info(
        f"주문 상세 페이지 로드 요청: id={dashboard_id}, user={current_user.get('user_id')}"
    )
    try:
        # 주문 데이터 조회
        order = get_dashboard_by_id(db, dashboard_id)
        if not order:
            logger.warning(f"주문 상세 로드 중 오류: 404, 주문을 찾을 수 없습니다.")
            raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다.")

        # 주문 데이터를 응답용 형식으로 변환
        # 객체를 딕셔너리로 변환하여 템플릿에서 접근 가능하게 함
        order_data = get_dashboard_response_data(order)

        # 필수 datetime 필드들 추가 (ISO 8601 문자열 형식)
        # create_time은 템플릿에서 필수적으로 사용되므로 항상 포함해야 함
        order_data["create_time"] = (
            order.create_time.isoformat() if order.create_time else None
        )
        order_data["depart_time"] = (
            order.depart_time.isoformat() if order.depart_time else None
        )
        order_data["complete_time"] = (
            order.complete_time.isoformat() if order.complete_time else None
        )
        # update_time이 응답 데이터에 없을 경우 추가
        if "update_time" not in order_data and hasattr(order, "update_at"):
            order_data["update_time"] = (
                order.update_at.isoformat() if order.update_at else None
            )
        # update_by가 응답 데이터에 없을 경우 추가
        if "update_by" not in order_data and hasattr(order, "update_by"):
            order_data["update_by"] = order.update_by

        # 누구나 수정 가능하므로 항상 editable=True 설정
        order_data["editable"] = True

        # 상세페이지에 필요한 메시지 파라미터 추출
        error_message = request.query_params.get("error")
        success_message = request.query_params.get("success")

        context = {
            "request": request,
            "current_user": current_user,
            "order": order_data,  # 딕셔너리 형태로 전달
            "error_message": error_message,
            "success_message": success_message,
            "current_user_role": current_user.get("user_role"),  # 추가: 권한 정보
        }
        return templates.TemplateResponse("order_page.html", context)

    except HTTPException as http_exc:
        logger.warning(
            f"주문 상세 로드 중 오류: {http_exc.status_code}, {http_exc.detail}"
        )
        error_message = quote(http_exc.detail)
        return RedirectResponse(
            url=f"/dashboard?error={error_message}",
            status_code=status.HTTP_303_SEE_OTHER,
        )
    except Exception as e:
        logger.error(f"주문 상세 페이지 로드 중 예외 발생: {e}", exc_info=True)
        error_message = quote("주문 정보를 불러오는 중 오류가 발생했습니다")
        return RedirectResponse(
            url=f"/dashboard?error={error_message}",
            status_code=status.HTTP_303_SEE_OTHER,
        )


# --- 주문 수정 페이지 --- (락 보유 확인 후 페이지 렌더링)
@page_router.get("/orders/{dashboard_id}/edit", name="order_edit_page")
@db_transaction  # 락 획득/해제 위해 트랜잭션 필요
async def order_edit_page(
    request: Request,
    dashboard_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """주문 수정 페이지 로드, 락 획득 후 렌더링"""
    user_id = current_user.get("user_id")
    logger.info(f"주문 수정 페이지 요청: id={dashboard_id}, user={user_id}")

    # 상세 페이지 URL (리다이렉트용)
    detail_url = f"/orders/{dashboard_id}"

    try:
        # 주문 정보 로드
        dashboard = get_dashboard_by_id(db, dashboard_id)
        if not dashboard:
            logger.warning(f"주문 수정 페이지 로드 실패: 주문을 찾을 수 없음")
            error_message = quote("수정할 주문을 찾을 수 없습니다.")
            return RedirectResponse(
                f"/dashboard?error={error_message}",
                status_code=status.HTTP_303_SEE_OTHER,
            )

        # 주문 정보 dict로 변환
        dashboard_data = get_dashboard_response_data(dashboard)

        # 필수 datetime 필드 추가
        dashboard_data["create_time"] = (
            dashboard.create_time.isoformat() if dashboard.create_time else None
        )
        dashboard_data["depart_time"] = (
            dashboard.depart_time.isoformat() if dashboard.depart_time else None
        )
        dashboard_data["complete_time"] = (
            dashboard.complete_time.isoformat() if dashboard.complete_time else None
        )

        # ETA ISO 8601 형식 변환 (YYYY-MM-DDTHH:MM)
        if dashboard.eta:
            dashboard_data["eta"] = dashboard.eta.isoformat()

        # 디버깅 메시지 추가
        logger.info(f"주문 수정 페이지 로드 성공: id={dashboard_id}")

        # 템플릿 컨텍스트 설정
        context = {
            "request": request,
            "order": dashboard_data,
            "current_user": current_user,
            "is_edit": True,
            "error_message": request.query_params.get("error"),
            "success_message": request.query_params.get("success"),
            "version": dashboard.version,  # 버전 정보 추가
        }

        return templates.TemplateResponse("order_form.html", context)

    except HTTPException as http_exc:
        # 표준 HTTP 예외 처리
        logger.warning(
            f"주문 수정 페이지 HTTP 예외: {http_exc.status_code}, {http_exc.detail}"
        )
        error_message = quote(http_exc.detail)
        return RedirectResponse(
            f"{detail_url}?error={error_message}",
            status_code=status.HTTP_303_SEE_OTHER,
        )
    except Exception as e:
        # 기타 예외 처리
        logger.error(f"주문 수정 페이지 로드 중 예외 발생: {str(e)}", exc_info=True)
        error_message = quote("페이지 로드 중 오류가 발생했습니다")
        return RedirectResponse(
            f"{detail_url}?error={error_message}",
            status_code=status.HTTP_303_SEE_OTHER,
        )


# --- 주문 수정 처리 API --- (Form 데이터 처리 및 서비스 호출)
@api_router.post("/orders/{dashboard_id}", status_code=status.HTTP_302_FOUND)
@db_transaction
async def update_order_action(
    request: Request,
    dashboard_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
    # Form 필드 (snake_case 사용)
    type: str = Form(...),
    department: str = Form(...),
    warehouse: str = Form(...),
    sla: str = Form(...),
    eta_str: str = Form(..., alias="eta"),
    postal_code: str = Form(...),
    address: str = Form(...),
    customer: str = Form(...),
    contact: Optional[str] = Form(None),
    remark: Optional[str] = Form(None),
    status_val: Optional[str] = Form(None, alias="status"),
    driver_name: Optional[str] = Form(None),
    driver_contact: Optional[str] = Form(None),
    delivery_company: Optional[str] = Form(None),
    version: int = Form(..., description="수정 대상의 현재 버전"),
):
    user_id = current_user.get("user_id")
    logger.info(f"주문 수정 API 요청: id={dashboard_id}, user={user_id}")

    # 리다이렉트 URL 설정
    detail_url = f"/orders/{dashboard_id}"
    edit_url = f"/orders/{dashboard_id}/edit"

    try:
        # 현재 DB 버전 확인
        current_order = get_dashboard_by_id(db, dashboard_id)
        if not current_order:
            # 서비스 함수에서도 동일한 체크가 있지만, 버전 비교 전에 확인하는 것이 좋음
            raise HTTPException(
                status_code=404, detail="수정할 주문을 찾을 수 없습니다."
            )

        version_mismatch = current_order.version != version
        if version_mismatch:
            logger.warning(
                f"주문 수정 버전 불일치: ID={dashboard_id}, Client Version={version}, DB Version={current_order.version}"
            )
            # 경고 메시지에 필요한 정보 (동시 수정 발생 시)
            concurrent_modifier = current_order.update_by
            concurrent_update_at = (
                current_order.update_at.strftime("%Y-%m-%d %H:%M")
                if current_order.update_at
                else "알 수 없음"
            )
            warning_msg = f"다른 사용자({concurrent_modifier})가 {concurrent_update_at}에 먼저 수정했습니다."
        else:
            warning_msg = None

        # ETA 파싱
        try:
            # ISO 8601 형식 처리 (YYYY-MM-DDTHH:MM)
            if "T" in eta_str:
                eta_dt = datetime.fromisoformat(eta_str)
            # 공백으로 구분된 경우 ISO 형식으로 변환
            else:
                eta_dt = datetime.fromisoformat(eta_str.replace(" ", "T"))
        except ValueError:
            logger.warning(f"ETA 형식 오류: {eta_str}")
            error_msg = "ETA 형식이 잘못되었습니다 (YYYY-MM-DDTHH:MM 형식이어야 합니다)"
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JSONResponse(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    content={"success": False, "message": error_msg},
                )
            else:
                return RedirectResponse(
                    f"{edit_url}?error={quote(error_msg)}",
                    status_code=status.HTTP_303_SEE_OTHER,
                )

        # 수정 데이터 준비
        update_data = {
            "type": type,
            "department": department,
            "warehouse": warehouse,
            "sla": sla,
            "eta": eta_dt,
            "postal_code": postal_code,
            "address": address,
            "customer": customer,
            "contact": contact,
            "remark": remark,
        }

        # 선택적 필드 추가
        if status_val:
            update_data["status"] = status_val
        if driver_name:
            update_data["driver_name"] = driver_name
        if driver_contact:
            update_data["driver_contact"] = driver_contact
        if delivery_company:
            update_data["delivery_company"] = delivery_company

        # 서비스 함수 호출
        result = update_dashboard(db, dashboard_id, update_data, user_id)

        # 성공 응답
        success_msg = "주문이 성공적으로 수정되었습니다."
        logger.info(f"주문 수정 성공: ID {dashboard_id}")

        # 리다이렉트 URL 생성 (warning 포함 가능)
        redirect_url = f"{detail_url}?success={quote(success_msg)}"
        if version_mismatch and warning_msg:  # 버전 불일치 시 경고 메시지 추가
            redirect_url += f"&warning={quote(warning_msg)}"

        # RedirectResponse 반환으로 복원
        return RedirectResponse(redirect_url, status_code=status.HTTP_303_SEE_OTHER)

    except HTTPException as http_exc:
        # HTTP 예외 발생 시 리다이렉트 (오류 메시지 포함)
        logger.warning(
            f"주문 수정 API HTTP 오류: {http_exc.status_code}, {http_exc.detail}"
        )
        # edit_url 로 리다이렉트하며 에러 표시 (기존 방식과 유사하게)
        error_message = quote(http_exc.detail)
        return RedirectResponse(
            f"{edit_url}?error={error_message}", status_code=status.HTTP_303_SEE_OTHER
        )

    except Exception as e:
        logger.error(f"주문 수정 API 처리 중 예외 발생: {e}", exc_info=True)
        error_message = quote("주문 수정 중 서버 오류가 발생했습니다.")
        # edit_url 로 리다이렉트하며 에러 표시
        return RedirectResponse(
            f"{edit_url}?error={error_message}", status_code=status.HTTP_303_SEE_OTHER
        )


# --- 주문 삭제 처리 API --- (서비스 호출 및 리다이렉트)
@api_router.post("/orders/{dashboard_id}/delete", status_code=status.HTTP_302_FOUND)
@db_transaction
async def delete_order_action(
    request: Request,
    dashboard_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_admin_user),  # ADMIN 전용
):
    user_id = current_user.get("user_id")
    user_role = current_user.get("user_role")
    logger.info(f"주문 삭제 API 요청: id={dashboard_id}, user={user_id}")

    detail_url = request.url_for("order_detail_page", dashboard_id=dashboard_id)
    dashboard_url = request.url_for("dashboard_page")

    try:
        # 서비스 함수 호출 - 내부에서 락 획득/확인 및 권한 체크, 삭제 후 락 해제
        result_list = delete_dashboard(
            db=db, dashboard_ids=[dashboard_id], user_id=user_id, user_role=user_role
        )

        # 결과 확인 및 리다이렉트
        if result_list and result_list[0].get("success"):
            # 성공 시 대시보드로 리다이렉트
            success_message = quote("주문이 성공적으로 삭제되었습니다.")
            return RedirectResponse(
                f"{dashboard_url}?success={success_message}",
                status_code=status.HTTP_303_SEE_OTHER,
            )
        else:
            # 서비스에서 실패 메시지를 반환한 경우
            error_message = (
                result_list[0].get("message", "삭제 실패")
                if result_list
                else "삭제 실패"
            )
            raise HTTPException(status_code=400, detail=error_message)

    except HTTPException as http_exc:
        logger.warning(
            f"주문 삭제 실패 (HTTPException): id={dashboard_id}, {http_exc.detail}"
        )

        # 락 오류(423)인 경우 특별 처리
        if http_exc.status_code == status.HTTP_423_LOCKED:
            error_message = quote(http_exc.detail)
            return RedirectResponse(
                f"{detail_url}?error={error_message}",
                status_code=status.HTTP_303_SEE_OTHER,
            )

        # 기타 HTTP 오류
        error_message = quote(http_exc.detail)
        return RedirectResponse(
            f"{detail_url}?error={error_message}", status_code=status.HTTP_303_SEE_OTHER
        )
    except Exception as e:
        logger.error(f"주문 삭제 API 처리 중 예외 발생: {e}", exc_info=True)
        error_message = quote("주문 삭제 중 오류가 발생했습니다")
        return RedirectResponse(
            f"{detail_url}?error={error_message}", status_code=status.HTTP_303_SEE_OTHER
        )


# --- 주문 생성 API ---
@api_router.post("/orders", status_code=status.HTTP_302_FOUND)
@db_transaction
async def create_order_action(
    request: Request,
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
    # Form 필드명 snake_case
    order_no: str = Form(...),
    type: str = Form(...),
    department: str = Form(...),
    warehouse: str = Form(...),
    sla: str = Form(...),
    eta_str: str = Form(..., alias="eta"),
    postal_code: str = Form(...),
    address: str = Form(...),
    customer: str = Form(...),
    contact: Optional[str] = Form(None),
    remark: Optional[str] = Form(None),
    status_val: Optional[str] = Form("WAITING", alias="status"),  # 기본값 WAITING
    driver_name: Optional[str] = Form(None),
    driver_contact: Optional[str] = Form(None),
    delivery_company: Optional[str] = Form(None),
):
    user_id = current_user.get("user_id")
    logger.info(f"주문 생성 API 요청: user={user_id}, order_no={order_no}")

    dashboard_url = request.url_for("dashboard_page")
    create_url = request.url_for("order_create_page")

    try:
        # ETA 파싱
        try:
            # ISO 8601 형식 처리 (YYYY-MM-DDTHH:MM)
            if "T" in eta_str:
                eta_dt = datetime.fromisoformat(eta_str)
            # 공백으로 구분된 경우 ISO 형식으로 변환
            else:
                eta_dt = datetime.fromisoformat(eta_str.replace(" ", "T"))
        except ValueError as e:
            logger.warning(f"ETA 형식 오류: {eta_str}, 오류: {str(e)}")
            error_message = quote(
                "ETA 형식이 잘못되었습니다 (YYYY-MM-DDTHH:MM 형식이어야 합니다)"
            )
            return RedirectResponse(
                f"{create_url}?error={error_message}",
                status_code=status.HTTP_303_SEE_OTHER,
            )

        # 생성 데이터 준비 (dict -> 모델)
        create_data = {
            "order_no": order_no,
            "type": type,
            "department": department,
            "warehouse": warehouse,
            "sla": sla,
            "eta": eta_dt,
            "postal_code": postal_code,
            "address": address,
            "customer": customer,
        }

        # 선택적 필드 추가
        if contact:
            create_data["contact"] = contact
        if remark:
            create_data["remark"] = remark
        if status_val and status_val != "WAITING":  # 기본값과 다를 경우만 추가
            create_data["status"] = status_val
        if driver_name:
            create_data["driver_name"] = driver_name
        if driver_contact:
            create_data["driver_contact"] = driver_contact
        if delivery_company:
            create_data["delivery_company"] = delivery_company

        # create_dashboard 서비스 호출 (Pydantic 모델 객체 생성)
        try:
            create_data_obj = DashboardCreate(**create_data)
        except Exception as e:
            logger.error(f"주문 생성 객체 생성 실패: {str(e)}", exc_info=True)
            error_message = quote(f"주문 데이터 형식 오류: {str(e)}")
            return RedirectResponse(
                f"{create_url}?error={error_message}",
                status_code=status.HTTP_303_SEE_OTHER,
            )

        # 서비스 호출하여 주문 생성
        new_dashboard = create_dashboard(db=db, data=create_data_obj, user_id=user_id)
        logger.info(
            f"주문 생성 성공: ID={new_dashboard.dashboard_id}, 주문번호={order_no}"
        )

        # 성공 시 새 주문 상세 페이지로 리다이렉트
        success_message = quote("새 주문이 성공적으로 생성되었습니다.")
        return RedirectResponse(
            f"/orders/{new_dashboard.dashboard_id}?success={success_message}",
            status_code=status.HTTP_303_SEE_OTHER,
        )

    except HTTPException as http_exc:
        # HTTP 예외 처리
        logger.warning(f"주문 생성 실패 (HTTPException): {http_exc.detail}")
        error_message = quote(http_exc.detail)
        return RedirectResponse(
            f"{create_url}?error={error_message}", status_code=status.HTTP_303_SEE_OTHER
        )
    except Exception as e:
        # 기타 예외
        logger.error(f"주문 생성 API 처리 중 예외 발생: {e}", exc_info=True)
        error_message = quote(
            "주문 생성 중 서버 오류가 발생했습니다. 관리자에게 문의하세요."
        )
        return RedirectResponse(
            f"{create_url}?error={error_message}", status_code=status.HTTP_303_SEE_OTHER
        )


# === 추가: 일괄 처리 지원 API ===


@api_router.get(
    "/orders/batch-get-common-statuses",
    summary="선택된 주문들의 공통 변경 가능 상태 조회",
)
async def get_batch_common_statuses(
    request: Request,
    ids: List[int] = Query(..., description="상태를 조회할 주문 ID 목록"),
    db: Session = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """주어진 주문 ID 목록에 대해 공통적으로 변경 가능한 다음 상태 목록을 반환합니다."""
    user_role = current_user.get("user_role")
    is_admin = user_role == "ADMIN"
    logger.info(
        f"공통 상태 조회 요청: IDs {ids}, User {current_user.get('user_id')}, Role {user_role}"
    )

    if not ids:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={
                "success": False,
                "message": "주문 ID가 제공되지 않았습니다.",
                "common_statuses": [],
                "labels": {},
            },
        )

    try:
        # 선택된 주문들의 현재 상태 조회
        orders = (
            db.query(Dashboard.status).filter(Dashboard.dashboard_id.in_(ids)).all()
        )

        if not orders:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "success": False,
                    "message": "선택된 주문을 찾을 수 없습니다.",
                    "common_statuses": [],
                    "labels": {},
                },
            )

        if len(orders) != len(ids):
            logger.warning(
                f"일부 주문 ID를 찾을 수 없음: 요청 {len(ids)}개, 찾음 {len(orders)}개"
            )
            # 일부만 찾은 경우 어떻게 처리할지 정책 필요 (여기서는 찾은 것만 기준으로 계산)

        current_statuses = [order.status for order in orders]

        # 각 주문별 다음 가능 상태 목록 계산
        possible_next_statuses_list = [
            set(get_next_possible_statuses(status, is_admin))
            for status in current_statuses
        ]

        # 모든 목록의 교집합 계산
        if not possible_next_statuses_list:  # 비어있는 경우 (이론상 발생 안 함)
            common_next_statuses = set()
        else:
            common_next_statuses = possible_next_statuses_list[0]
            for status_set in possible_next_statuses_list[1:]:
                common_next_statuses.intersection_update(status_set)

        # 현재 상태들도 결과에 포함할지 여부 결정 (선택)
        # 예: 현재 WAITING 주문과 IN_PROGRESS 주문을 동시에 선택 시 공통 다음 상태는 ISSUE, CANCEL 뿐
        # 사용성 측면에서 현재 상태들을 포함시켜주는 것이 좋을 수 있음. (단, 모든 주문이 동일 상태일 때만 의미있음)
        # 여기서는 교집합 결과만 반환

        common_statuses_list = sorted(list(common_next_statuses))  # 정렬하여 반환

        # 상태 라벨 정보도 함께 반환
        common_status_labels = {
            status: status_labels.get(status, status) for status in common_statuses_list
        }

        logger.info(f"계산된 공통 상태: {common_statuses_list}")

        return {
            "success": True,
            "common_statuses": common_statuses_list,
            "labels": common_status_labels,
        }

    except Exception as e:
        logger.error(f"공통 상태 조회 중 서버 오류: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="공통 상태 조회 중 서버 오류 발생",
        )
